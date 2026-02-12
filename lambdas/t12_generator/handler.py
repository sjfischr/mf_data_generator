"""T-12 and Rent Roll Generator Lambda — creates Excel files."""

import io
import json
import logging
import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lambdas.shared import bedrock_client, s3_utils
from lambdas.shared.models import CrosswalkData

logger = logging.getLogger(__name__)

FIRST_NAMES = [
    "James", "Mary", "Robert", "Patricia", "John", "Jennifer", "Michael",
    "Linda", "David", "Elizabeth", "William", "Barbara", "Richard", "Susan",
    "Joseph", "Jessica", "Thomas", "Sarah", "Christopher", "Karen", "Charles",
    "Lisa", "Daniel", "Nancy", "Matthew", "Betty", "Anthony", "Margaret",
    "Mark", "Sandra", "Donald", "Ashley", "Steven", "Dorothy", "Andrew",
    "Kimberly", "Paul", "Emily", "Joshua", "Donna",
]

LAST_NAMES = [
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
    "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
    "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin",
    "Lee", "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark",
    "Ramirez", "Lewis", "Robinson",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
CURRENCY_FMT = '#,##0'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def generate_rent_roll(crosswalk: CrosswalkData) -> bytes:
    """Generate a unit-by-unit rent roll as XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rent Roll"

    prop = crosswalk.property_identification
    fin = crosswalk.financial_data
    phys = crosswalk.property_physical

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Rent Roll — {prop.property_name}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells("A2:I2")
    ws["A2"] = f"As of {fin.effective_date}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)

    # Headers
    headers = [
        "Unit #", "Unit Type", "SF", "Tenant Name", "Lease Start",
        "Lease End", "Monthly Rent", "Market Rent", "Status",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Generate unit rows
    row = 5
    unit_num = 100
    effective = date(2026, 2, 10)
    occupied_count = fin.occupancy.occupied_units
    total_generated = 0

    for unit_item in phys.unit_mix:
        unit_type = unit_item.unit_type
        market_rent = fin.market_rents_monthly.get(unit_type, 1400)
        in_place_rent = fin.in_place_rents_monthly.get(unit_type, 1350)

        for i in range(unit_item.count):
            is_occupied = total_generated < occupied_count
            unit_number = f"{unit_num + i}"

            if is_occupied:
                tenant = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
                lease_start = effective - timedelta(days=random.randint(30, 365))
                lease_end = lease_start + timedelta(days=365)
                rent = in_place_rent + random.randint(-25, 25)
                status = "Occupied"
            else:
                tenant = "— Vacant —"
                lease_start = None
                lease_end = None
                rent = 0
                status = "Vacant"

            ws.cell(row=row, column=1, value=unit_number)
            ws.cell(row=row, column=2, value=unit_type)
            ws.cell(row=row, column=3, value=unit_item.avg_size_sf)
            ws.cell(row=row, column=4, value=tenant)
            ws.cell(row=row, column=5, value=lease_start.isoformat() if lease_start else "")
            ws.cell(row=row, column=6, value=lease_end.isoformat() if lease_end else "")
            ws.cell(row=row, column=7, value=rent).number_format = CURRENCY_FMT
            ws.cell(row=row, column=8, value=market_rent).number_format = CURRENCY_FMT
            ws.cell(row=row, column=9, value=status)

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).font = DATA_FONT

            row += 1
            total_generated += 1

        unit_num += unit_item.count + 10

    # Column widths
    widths = [8, 12, 8, 22, 12, 12, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_t12(crosswalk: CrosswalkData, year_key: str, year_label: str) -> bytes:
    """Generate a T-12 operating statement as XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"T-12 {year_label}"

    prop = crosswalk.property_identification
    hist = crosswalk.financial_data.historical_t12[year_key]

    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = f"Trailing 12-Month Operating Statement — {prop.property_name}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells("A2:N2")
    ws["A2"] = f"{year_label}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)

    # Headers: Category + 12 months + Annual Total
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers = ["Category"] + months + ["Annual Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Distribute annual values across months with slight variation
    def monthly_spread(annual: int) -> list[int]:
        base = annual // 12
        values = [base + random.randint(-int(base * 0.05), int(base * 0.05)) for _ in range(11)]
        values.append(annual - sum(values))
        return values

    # Income section
    categories = [
        ("INCOME", None),
        ("Rental Income", hist.rental_income),
        ("Other Income", hist.other_income),
        ("Less: Vacancy Loss", hist.vacancy_loss),
        ("Effective Gross Income", hist.effective_gross_income),
        ("", None),
        ("EXPENSES", None),
        ("Operating Expenses", hist.operating_expenses),
        ("", None),
        ("NET OPERATING INCOME", hist.net_operating_income),
    ]

    row = 5
    for label, annual in categories:
        ws.cell(row=row, column=1, value=label).font = DATA_FONT

        if label in ("INCOME", "EXPENSES", ""):
            ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True)
        elif annual is not None:
            monthly = monthly_spread(annual)
            for col, val in enumerate(monthly, 2):
                ws.cell(row=row, column=col, value=val).number_format = CURRENCY_FMT
                ws.cell(row=row, column=col).font = DATA_FONT
            ws.cell(row=row, column=14, value=annual).number_format = CURRENCY_FMT
            ws.cell(row=row, column=14).font = Font(name="Calibri", size=11, bold=True)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

        if label == "NET OPERATING INCOME":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True)

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 24
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def handler(event, context):
    job_id = event["job_id"]
    logger.info("Generating T-12 and Rent Roll for job %s", job_id)

    data = s3_utils.read_json(job_id, "crosswalk-data.json")
    crosswalk = CrosswalkData.model_validate(data)

    # Generate rent roll
    rent_roll_bytes = generate_rent_roll(crosswalk)
    s3_utils.write_bytes(
        job_id, "outputs/rent_roll.xlsx", rent_roll_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Generate T-12s for each historical year
    for year_key, year_label in [
        ("Year 1", "Year 1"),
        ("Year 2", "Year 2"),
        ("Year 3", "Year 3"),
    ]:
        t12_bytes = generate_t12(crosswalk, year_key, year_label)
        s3_utils.write_bytes(
            job_id, f"outputs/t12_{year_key}.xlsx", t12_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    logger.info("T-12 and Rent Roll generation complete for job %s", job_id)

    return {
        "status": "success",
        "job_id": job_id,
        "files": [
            "outputs/rent_roll.xlsx",
            "outputs/t12_year1.xlsx",
            "outputs/t12_year2.xlsx",
            "outputs/t12_year3.xlsx",
        ],
    }
